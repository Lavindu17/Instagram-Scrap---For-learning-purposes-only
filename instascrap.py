import instaloader
import time
import random
from time import sleep
import os
import sys
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Initialize Instaloader instance with more conservative settings
L = instaloader.Instaloader(
    download_pictures=False,
    download_videos=False,
    download_video_thumbnails=False,
    download_geotags=False,
    download_comments=False,
    save_metadata=False,
    compress_json=False,
    max_connection_attempts=3,
    # Rotate user agents to avoid detection
    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    # Add request timeout to prevent hanging
    request_timeout=30
)

# Function to log in to Instagram
def login(username, password, max_attempts=3):
    """
    Log in to Instagram with retry mechanism and session management.
    """
    for attempt in range(max_attempts):
        try:
            print(f"Logging in to Instagram (attempt {attempt+1}/{max_attempts})...")
            
            # First try to load an existing session
            try:
                L.load_session_from_file(username)
                print("Session loaded successfully!")
                
                # Verify if the session is still valid by checking profile
                L.check_profile_id(username)
                print("Session is valid.")
                return True
            except (FileNotFoundError, instaloader.exceptions.InvalidSessionException):
                print("No valid session found. Logging in with credentials...")
            
            # Login with username and password
            L.login(username, password)
            L.save_session_to_file(username)
            print("Login successful!")
            return True
            
        except instaloader.exceptions.BadCredentialsException:
            print("Error: Invalid username or password.")
            return False
        except instaloader.exceptions.ConnectionException as e:
            print(f"Connection error: {e}")
            if attempt < max_attempts - 1:
                wait_time = 30 * (attempt + 1)
                print(f"Retrying in {wait_time} seconds...")
                sleep(wait_time)
            else:
                print("Max login attempts reached.")
                return False
        except instaloader.exceptions.TwoFactorAuthRequiredException:
            print("Two-factor authentication is required.")
            two_factor_code = input("Enter the 2FA code sent to your device: ")
            try:
                L.two_factor_login(two_factor_code)
                L.save_session_to_file(username)
                print("2FA login successful!")
                return True
            except Exception as e:
                print(f"2FA login failed: {e}")
                return False
        except Exception as e:
            print(f"Unexpected error during login: {e}")
            if attempt < max_attempts - 1:
                wait_time = 30 * (attempt + 1)
                print(f"Retrying in {wait_time} seconds...")
                sleep(wait_time)
            else:
                print("Max login attempts reached.")
                return False
    
    return False

# Function to extract shortcode from URL
def get_shortcode(url):
    """Extract the post shortcode from a URL."""
    # Clean URL by removing trailing slashes
    url = url.rstrip('/')
    
    # Handle different URL formats
    if '/p/' in url:
        parts = url.split('/p/')
        if len(parts) > 1:
            shortcode = parts[1].split('/')[0]
            return shortcode
    
    print("Invalid Instagram post URL format. Expected format: https://www.instagram.com/p/SHORTCODE/")
    return None

# Smart rate limiter function with more conservative timing
def smart_sleep(idx=None, force_sleep=False):
    """
    Implements a smart rate limiting strategy to avoid Instagram restrictions.
    Using more conservative timings to avoid rate limiting.
    """
    # Base delay for all requests (3-5 seconds instead of 1.5-3)
    base_delay = random.uniform(3, 5)
    
    # Force sleep or every 5 requests: medium pause (8-12 seconds)
    if force_sleep or (idx is not None and idx % 5 == 0 and idx > 0):
        delay = random.uniform(8, 12)
        print(f"Taking a short break ({delay:.1f}s) to avoid rate limits...")
    # Every 20 requests: longer pause (25-40 seconds)
    elif idx is not None and idx % 20 == 0 and idx > 0:
        delay = random.uniform(25, 40)
        print(f"Taking a longer break ({delay:.1f}s) to avoid rate limits...")
    else:
        delay = base_delay
    
    sleep(delay)
    return

# Get post likes using a more reliable method - using post.get_likes()
def get_post_likes(post, max_likes=None):
    """Get users who liked the post using a more reliable approach."""
    likers = []
    like_count = 0
    
    try:
        print("\nFetching likes (this may take some time)...")
        print("To avoid Instagram rate limiting, we'll proceed slowly...")
        
        # Loop through post likes with smart rate limiting
        for idx, profile in enumerate(post.get_likes(), 1):
            username = profile.username
            profile_url = f"https://www.instagram.com/{username}/"
            
            print(f"[{idx}] Liked by: {username}")
            likers.append({
                'username': username,
                'profile_url': profile_url,
                'type': 'like'
            })
            like_count += 1
            
            # Apply more conservative rate limiting
            smart_sleep(idx)
            
            # Stop if we've reached max_likes
            if max_likes and idx >= max_likes:
                print(f"Reached maximum number of likes to fetch ({max_likes}).")
                break
    
    except instaloader.exceptions.LoginRequiredException:
        print("Warning: Cannot fetch likes. Login required.")
    except instaloader.exceptions.TooManyRequestsException:
        print("Instagram is rate limiting us. Wait a while before trying again.")
        print(f"We were able to retrieve {like_count} likes before being rate limited.")
    except Exception as e:
        print(f"Error fetching likes: {e}")
        print(f"We were able to retrieve {like_count} likes before the error.")
    
    return likers

# Function to safely get comments
def get_post_comments(post, max_comments=None):
    """Get list of users who commented on the post."""
    commenters = []
    try:
        print("\nFetching comments...")
        
        for idx, comment in enumerate(post.get_comments(), 1):
            username = comment.owner.username
            profile_url = f"https://www.instagram.com/{username}/"
            comment_text = comment.text
            
            print(f"[{idx}] Comment by: {username}")
            commenters.append({
                'username': username,
                'profile_url': profile_url,
                'comment': comment_text,
                'type': 'comment'
            })
            
            # Apply smart rate limiting
            smart_sleep(idx)
            
            # Stop if we've reached max_comments
            if max_comments and idx >= max_comments:
                print(f"Reached maximum number of comments to fetch ({max_comments}).")
                break
                
    except Exception as e:
        print(f"Error fetching comments: {e}")
    
    return commenters

# Function to get post details with alternative likes fetching method
def get_post_details(url, output_file=None, max_likes=100, max_comments=100):
    """
    Fetch details of an Instagram post including likes and comments.
    Optionally save results to an Excel file.
    Using a simplified approach to avoid rate limiting.
    """
    shortcode = get_shortcode(url)
    if not shortcode:
        return
    
    print(f"Fetching post with shortcode: {shortcode}")
    
    # Create lists to store data
    interactions = []
    
    retry_count = 0
    max_retries = 5
    post_info = {}
    
    while retry_count < max_retries:
        try:
            # Force a small delay before fetching the post
            smart_sleep(force_sleep=True)
            
            post = instaloader.Post.from_shortcode(L.context, shortcode)
            
            # Basic post info
            post_info = {
                'shortcode': post.shortcode,
                'owner_username': post.owner_username,
                'owner_profile_url': f"https://www.instagram.com/{post.owner_username}/",
                'date': post.date,
                'caption': post.caption or "",
                'likes_count': post.likes,
                'comments_count': post.comments,
                'url': f"https://www.instagram.com/p/{post.shortcode}/"
            }
            
            # Show basic post info
            print(f"\nPost by: {post_info['owner_username']}")
            print(f"Posted on: {post_info['date'].strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"Likes: {post_info['likes_count']}, Comments: {post_info['comments_count']}")
            
            # Get comments first (more reliable than likes)
            print("\nCollecting comments first...")
            commenters = get_post_comments(post, max_comments)
            interactions.extend(commenters)
            
            # Take a longer break between comments and likes
            sleep(random.uniform(15, 20))
            print("Taking a short break before fetching likes...")
            
            # Get likes with improved method
            likers = get_post_likes(post, max_likes)
            interactions.extend(likers)
            
            # Save results to Excel file
            if output_file:
                save_data_to_excel(output_file, post_info, interactions)
            
            print("\nData collection completed.")
            return post_info, interactions
            
        except instaloader.exceptions.InstaloaderException as e:
            retry_count += 1
            wait_time = 60 * retry_count
            print(f"Instagram error: {e}")
            if retry_count < max_retries:
                print(f"Retrying in {wait_time} seconds... (Attempt {retry_count}/{max_retries})")
                sleep(wait_time)
            else:
                print("Max retries reached. Could not fetch the post details.")
        except Exception as e:
            print(f"Unexpected error: {e}")
            break
    
    # If we got here with populated post_info but no interactions, still save what we have
    if post_info and output_file:
        save_data_to_excel(output_file, post_info, interactions)
    
    return post_info, interactions

# Simplified Excel export with all data in one sheet
def save_data_to_excel(filename, post_info, interactions):
    """Save data to a simplified Excel file with one sheet for all interactions."""
    try:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        timestamped_filename = f"{filename.rsplit('.', 1)[0]}_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram Data"
        
        # Create header section
        ws['A1'] = "Instagram Post Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Post information
        ws['A3'] = "Post URL:"
        ws['B3'] = post_info['url']
        ws['B3'].hyperlink = post_info['url']
        ws['B3'].font = Font(color="0563C1", underline="single")
        ws.merge_cells('B3:F3')
        
        ws['A4'] = "Posted by:"
        ws['B4'] = post_info['owner_username']
        ws['C4'] = post_info['owner_profile_url']
        ws['C4'].hyperlink = post_info['owner_profile_url']
        ws['C4'].font = Font(color="0563C1", underline="single")
        
        ws['A5'] = "Posted on:"
        ws['B5'] = post_info['date'].strftime('%Y-%m-%d %H:%M:%S')
        
        ws['A6'] = "Total Likes:"
        ws['B6'] = post_info['likes_count']
        
        ws['A7'] = "Total Comments:"
        ws['B7'] = post_info['comments_count']
        
        # Count collected interactions
        likes_collected = sum(1 for item in interactions if item.get('type') == 'like')
        comments_collected = sum(1 for item in interactions if item.get('type') == 'comment')
        
        ws['A8'] = "Likes Collected:"
        ws['B8'] = likes_collected
        
        ws['A9'] = "Comments Collected:"
        ws['B9'] = comments_collected
        
        # Separator
        ws['A11'] = "INTERACTIONS (LIKES AND COMMENTS)"
        ws['A11'].font = Font(bold=True)
        ws.merge_cells('A11:F11')
        ws['A11'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Headers for the interaction data
        headers = ["#", "Type", "Username", "Profile URL", "Comment/Timestamp"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Populate data
        for idx, interaction in enumerate(interactions, 1):
            row = idx + 12
            
            # Index
            ws.cell(row=row, column=1, value=idx)
            
            # Type (like or comment)
            ws.cell(row=row, column=2, value=interaction.get('type', 'unknown'))
            
            # Username
            ws.cell(row=row, column=3, value=interaction['username'])
            
            # Profile URL as hyperlink
            cell = ws.cell(row=row, column=4, value=interaction['profile_url'])
            cell.hyperlink = interaction['profile_url']
            cell.font = Font(color="0563C1", underline="single")
            
            # Comment text or timestamp (for likes)
            if interaction.get('type') == 'comment':
                ws.cell(row=row, column=5, value=interaction.get('comment', ''))
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 50
        
        # Save the workbook
        wb.save(timestamped_filename)
        print(f"Results saved to Excel file: {timestamped_filename}")
    except Exception as e:
        print(f"Error saving results to Excel file: {e}")

def main():
    print("Instagram Post Analysis Tool")
    print("===========================")
    
    # Input credentials
    username = input("Enter your Instagram username: ")
    password = input("Enter your Instagram password: ")
    
    # Attempt login
    if not login(username, password):
        print("Login failed. Exiting.")
        return
    
    while True:
        # Input the Instagram post URL
        url = input("\nEnter the Instagram post URL (or 'exit' to quit): ")
        if url.lower() == 'exit':
            break
        
        # Ask for maximum likes/comments to fetch
        try:
            max_likes = int(input("Maximum number of likes to fetch (default: 100, enter 0 for all): ") or "100")
            if max_likes <= 0:
                max_likes = None  # Fetch all
        except ValueError:
            max_likes = 100
            
        try:
            max_comments = int(input("Maximum number of comments to fetch (default: 100, enter 0 for all): ") or "100")
            if max_comments <= 0:
                max_comments = None  # Fetch all
        except ValueError:
            max_comments = 100
        
        # Always save to Excel
        output_file = f"instagram_data_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Get post details
        get_post_details(url, output_file, max_likes, max_comments)
        
        # Ask if user wants to analyze another post
        continue_option = input("\nWould you like to analyze another post? (y/n): ").lower()
        if continue_option != 'y':
            break
    
    print("Thank you for using Instagram Post Analysis Tool!")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProcess interrupted by user. Exiting...")
    except Exception as e:
        print(f"Unexpected error: {e}")
